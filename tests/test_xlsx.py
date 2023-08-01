import os
from openpyxl import load_workbook
from .conftest import resources_path


# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_by_xlsx():
    workbook = load_workbook(os.path.join(resources_path, 'file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    n = sheet.cell(row=3, column=2).value
    print(n)

    assert 'Teresa' in n

